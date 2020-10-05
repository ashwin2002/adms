import openpyxl
import os
import traceback

from lib import commonLib


class TextToExcel:
    def __init__(self, progress_bar, log_viewer):
        self.progressBar = progress_bar
        self.logViewer = log_viewer
        return

    def logMsg (self, level, message):
        if level == 'ERROR':
            message = message.replace ('\n', '<br/>')
            (self.logViewer).append ('<font color="red">%s: %s</font>' % (level, message))
        elif level == 'WARNING':
            message = message.replace ('\n', '<br/>')
            (self.logViewer).append ('<font color="#e6ac00">%s: %s</font>' % (level, message))
        else:
            (self.logViewer).append ('%s: %s' % (level, message))
        return

    def checkFileExists (self, inputFileName, printError=True):
        errorCondition = False
        if not(os.path.isfile (inputFileName)):
            if printError:
                message = "File '%s' does not exists !" % (inputFileName)
                self.logMsg ('ERROR', message)
            errorCondition = True
        return errorCondition

    def appendRowToWorksheet(self, worksheetObj, dataRow):
        worksheetObj.append(dataRow)
        return

    def convert(self, input_file_name):
        self.logMsg('INFO', "Processing file '%s'" % input_file_name)
        output_file_name = input_file_name.split('/')
        worksheet = workbook = None
        file_name = (output_file_name.pop()).split('.')[0]
        if len(file_name) != 8:
            self.logMsg('ERROR', "Invalid file name format '%s'. Use XXYY9999 format" % file_name)
        else:
            comp_str = file_name[0:2]
            month_num = int(file_name[4:6])
            yr_num = int(file_name[6:8])
            data_type_str = file_name[2:4]
            output_file_name = '/'.join(output_file_name)
            output_file_name += '/%s_%s%d.xlsx' % (comp_str, commonLib.getMonthStr(month_num, 3), yr_num)

            fileNotExists = self.checkFileExists(output_file_name, printError=False)
            if fileNotExists:
                workbook = openpyxl.Workbook()
                workbook.guess_types = True
                worksheet = workbook.create_sheet(data_type_str)
            else:
                workbook = openpyxl.load_workbook(output_file_name)
                worksheet_list = workbook.get_sheet_names()

                if 'Sheet' in worksheet_list:
                    worksheet = workbook.get_sheet_by_name('Sheet')
                    workbook.remove_sheet(worksheet)

                if data_type_str in worksheet_list:
                    worksheet = workbook.get_sheet_by_name(data_type_str)
                    workbook.remove_sheet(worksheet)

                worksheet = workbook.create_sheet(data_type_str)

        col_data_dict = None
        try:
            with open(input_file_name, 'r') as filePtr:
                while True:
                    txt_file_line = filePtr.readline()
                    if not txt_file_line:
                        break

                    if txt_file_line.strip() == '':
                        continue

                    if not col_data_dict:
                        col_data_dict = dict()
                        char_index = 0
                        in_word = False
                        curr_word_index = -1
                        for char in txt_file_line:
                            if char == ' ' and in_word:
                                col_data_dict[curr_word_index] = char_index
                                in_word = False
                            elif char != ' ' and not in_word:
                                in_word = True
                                curr_word_index = char_index
                                col_data_dict[char_index] = -1
                            char_index += 1
                        sorted_dict_keys = sorted(col_data_dict.keys())
                    else:
                        line_len = len(txt_file_line)
                        xl_data_row = list()

                        for field_index, startIndex \
                                in enumerate(sorted_dict_keys):
                            if line_len < startIndex:
                                continue

                            data = txt_file_line[startIndex:col_data_dict[startIndex]].strip()
                            try:
                                if field_index != 3:
                                    data = float(data)
                                else:
                                    try:
                                        float(data)
                                        data = int(data)
                                    except:
                                        pass
                                    finally:
                                        data = "=CONCATENATE(\"%s\")" \
                                               % int(data)
                            except:
                                pass

                            xl_data_row.append(data)

                        self.appendRowToWorksheet(worksheet, xl_data_row)

            workbook.save(output_file_name)
        except Exception:
            self.logMsg ('ERROR', 'Exception during file parsing -> %s' % (traceback.format_exc()))

        self.logMsg ('INFO', "Output saved in file '%s'" % output_file_name)
        self.logMsg ('::::::::::', 'Done ::::::::::')
