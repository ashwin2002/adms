import json
import openpyxl
import os
import traceback

from lib import commonLib
from lib.excel_lib import is_file_exists
from lib.logging import UiLogger


def get_value_from_dict(dict_obj, field_name, field_type):
    value = ''
    if field_type == 'amt' or field_type == 'rate':
        value = 0.00

    if field_name in dict_obj:
        value = dict_obj[field_name]
    return value


class JsonToExcel:
    def __init__(self):
        self.log = None

    def loadJsonData(self, json_file_name):
        with open(json_file_name, 'r') as filePtr:
            json_obj = json.loads(filePtr.read())
        return json_obj

    def print_gstin(self, json_obj):
        if 'gstin' in json_obj:
            gstin_in_file = json_obj['gstin']
            self.log.info("GSTIN as per file '%s'" % gstin_in_file)
        else:
            self.log.info("GSTIN not present in json data")

    @staticmethod
    def get_value_for_key(key, lookup_dict):
        value = key + ' not available'
        if key in lookup_dict:
            value = lookup_dict[key]
        return value


class Gstr1(JsonToExcel):
    def __init__(self, progress_bar, log_viewer):
        self.progressBar = progress_bar
        self.log = UiLogger(log_viewer)

    def writeHsnData(self, jsonObj, worksheetObj):
        if 'hsn' in jsonObj:
            hsnDict = jsonObj['hsn']
            if 'data' in hsnDict:
                self.log.info('Writing HSN related data')
                hsnHeadingDataRow = ['HSN Code', 'Product Desc', 'Units', 'S.Amt', 'C.Amt', 'Quantity', 'Value', 'Tax Value', 'Number', 'CS Amt', 'I.Amt']
                worksheetObj.append(hsnHeadingDataRow)

                for hsnDataRow in hsnDict['data']:
                    dataDesc = get_value_from_dict(hsnDataRow, 'desc', 'str')
                    for field in ['csamt', 'iamt', 'uqc']:
                        if field not in hsnDataRow:
                            hsnDataRow[field] = 0
                    hsnDataRow = [hsnDataRow['hsn_sc'], dataDesc, hsnDataRow['uqc'], hsnDataRow['samt'], hsnDataRow['camt'], hsnDataRow['qty'],
                                  hsnDataRow['val'], hsnDataRow['txval'], hsnDataRow['num'], hsnDataRow['csamt'], hsnDataRow['iamt']]
                    worksheetObj.append(hsnDataRow)
            else:
                self.log.error('HSN dictionary has no data array information')
        else:
            self.log.warning('Input file does not contain HSN related data')
        return

    def writeB2CSData(self, stateList, jsonObj, worksheetObj):
        if 'b2cs' in jsonObj:
            self.log.info('Writing B2CS related data')
            b2csHeadingRow = ['CS Amt', 'S Amt', 'Rate', 'Place of supply', 'Tax Value', 'Type', 'C Amt', 'Supply Type']
            worksheetObj.append(b2csHeadingRow)

            for b2csRow in jsonObj['b2cs']:
                posVal = int(b2csRow['pos'])
                posData = '%02d - %s' % (posVal, stateList[posVal - 1])

                csamt = get_value_from_dict(b2csRow, 'csamt', 'amt')
                samt = get_value_from_dict(b2csRow, 'samt', 'amt')
                rt = get_value_from_dict(b2csRow, 'rt', 'rate')
                txVal = get_value_from_dict(b2csRow, 'txval', 'rate')
                typ = get_value_from_dict(b2csRow, 'typ', 'str')
                camt = get_value_from_dict(b2csRow, 'camt', 'amt')
                splyType = get_value_from_dict(b2csRow, 'sply_ty', 'str')
                b2csDataRow = [csamt, samt, rt, posData, txVal, typ, camt, splyType]
                worksheetObj.append(b2csDataRow)
        else:
            self.log.info('Input file does not contain B2CS related data')
        return

    def write_b2b_data(self, stateList, jsonObj, worksheetObj):
        if 'b2b' in jsonObj:
            self.log.info('Writing B2B related data')
            b2bHeadingRow = ['GSTIN / UIN of Recipient', 'Invoice Number', 'Invoice Date', 'Invoice Value', 'Place of supply', 'Reverse Charge',
                             'Invoice Type', 'E-Commerce Gstin', 'Rate', 'Taxable Value', 'Cess Amount']
            worksheetObj.append(b2bHeadingRow)
            for b2bRow in jsonObj['b2b']:
                ctin = b2bRow['ctin']
                if 'inv' in b2bRow:
                    for invData in b2bRow['inv']:
                        invNum  = invData['inum']
                        invDate = invData['idt']
                        invVal  = invData['val']
                        invType = invData['inv_typ']
                        invPos  = invData['pos']
                        invEcomGstin = ''
                        invRevCharge = invData['rchrg']

                        try:
                            posVal = int(invData['pos'])
                            invPos = '%02d - %s' % (posVal, stateList[posVal - 1])
                        except:
                            invPos = 'State not available'

                        if invType == 'R':
                            invType = 'Regular'

                        if 'itms' in invData:
                            for itemData in invData['itms']:
                                if 'itm_det' in itemData:
                                    invTaxableVal  = itemData['itm_det']['txval']
                                    invTaxRate = itemData['itm_det']['rt']
                                    invCessAmt = ''
                                    b2bDataRow = [ctin, invNum, invDate, invVal, invPos, invRevCharge, invType, invEcomGstin, invTaxRate, invTaxableVal, invCessAmt]
                                    worksheetObj.append(b2bDataRow)
                                else:
                                    self.log.error("No item description available for invoice number '%s' dated '%s'" % (invNum, invDate))
                        else:
                            self.log.error("No item data present for invoice '%s' for customer '%s'" % (invNum, ctin))
                else:
                    self.log.error("No Invoice data present for '%s'" % ctin)
        else:
            self.log.warning('Input file does not contain B2B related data')
        return

    def writeErrData(self, stateList, jsonObj, worksheetObj):
        errTableHeadingRow = ['Invoice #', 'Invoice Date', 'Customer TIN', 'Error Code', 'Error Message',
                              'Invoice Type', 'Place of Sale', 'No. of Items', 'S.Amt', 'C.Amt', 'Tax Rate', 'Taxable Amt', 'Invoice Value', 'Reverse Charge']
        if 'error_report' in jsonObj:
            for (errHeading, errData) in jsonObj['error_report'].items():
                self.log.info('Writing %s error report' % errHeading)
                errHeading = 'Error Report for %s' % errHeading
                currErrorHeading = [errHeading]
                worksheetObj.append(currErrorHeading)
                worksheetObj.append(errTableHeadingRow)
                for errRow in errData:
                    custTin = self.get_value_for_key('ctin', errRow)
                    errMsg  = self.get_value_for_key('error_msg', errRow)
                    errCode = self.get_value_for_key('error_cd', errRow)
                    if 'inv' in errRow:
                        for invData in errRow['inv']:
                            invVal  = self.get_value_for_key('val', invData)
                            invType = self.get_value_for_key('inv_typ', invData)
                            try:
                                invPos = int(self.get_value_for_key('pos', invData))
                            except:
                                invPos = -1
                            invNum  = self.get_value_for_key('inum', invData)
                            invDate = self.get_value_for_key('idt', invData)
                            invRevCharge = self.get_value_for_key('rchrg', invData)

                            invItems = 'NA'
                            invItemDesc = 'NA'
                            invItemNum  = 'NA'
                            invItemSAmt = 'NA'
                            invItemCAmt = 'NA'
                            invItemTaxRate = 'NA'
                            invItemTaxableVal = 'NA'

                            if 'itms' in invData:
                                invItems = invData['itms'][0]

                                if 'itm_det' in invItems:
                                    invItemDesc = invItems['itm_det']

                                    if 'samt' in invItemDesc:
                                        invItemSAmt = invItemDesc['samt']
                                    if 'camt' in invItemDesc:
                                        invItemCAmt = invItemDesc['camt']
                                    if 'rt' in invItemDesc:
                                        invItemTaxRate = invItemDesc['rt']
                                    if 'txval' in invItemDesc:
                                        invItemTaxableVal = invItemDesc['txval']

                                if 'num' in invItems:
                                    invItemNum  = invItems['num']
                            # else:
                            #     self.log.error("Item data not available in JSON for inv# '%s' dated '%s'" % (invNum, invDate))

                            if invType == 'R':
                                invType = 'Regular'

                            invPos = '%02d - %s' % (invPos, stateList[invPos - 1])

                            errDataRow = [invNum, invDate, custTin, errCode, errMsg,
                                          invType, invPos, invItemNum, invItemSAmt, invItemCAmt, invItemTaxRate, invItemTaxableVal, invVal, invRevCharge]
                            worksheetObj.append(errDataRow)
                    else:
                        self.log.error("No invoice data available for '%s'"
                                       % custTin)
                worksheetObj.append([])
                worksheetObj.append([])
        else:
            self.log.warning('Input file does not contain Error_Report data')
        return

    def convert(self, jsonFileName):
        outputFileName = jsonFileName.replace('.json', '.xlsx')
        stateList = commonLib.getStateList()
        # Check input file exists in file system
        self.log.info('Checking input file exists or not')
        if is_file_exists(jsonFileName):
            return

        # Load JSON data from file
        self.log.info("Loading JSON data from '%s'" % jsonFileName)
        json_obj = self.loadJsonData(jsonFileName)
        if not json_obj:
            self.log.error('Unable to read JSON data from file')
            return

        self.print_gstin(json_obj)

        # Open Excel file and create required work sheets
        workbook = openpyxl.Workbook()
        hsnWorksheet = workbook.create_sheet('HSN')
        b2csWorksheet = workbook.create_sheet('B2CS')
        b2bWorksheet = workbook.create_sheet('B2B')
        errWorksheet = workbook.create_sheet('Error Report')

        try:
            self.writeHsnData(json_obj, hsnWorksheet)
            self.writeB2CSData(stateList, json_obj, b2csWorksheet)
            self.write_b2b_data(stateList, json_obj, b2bWorksheet)
            self.writeErrData(stateList, json_obj, errWorksheet)
        except Exception:
            self.log.error('Exception during file parsing -> %s'
                           % traceback.format_exc())

        workbook.save(outputFileName)
        self.log.info("Output saved in file '%s'" % outputFileName)
        self.log.raw_line(':::::::::: Done ::::::::::')


class Gstr2(JsonToExcel):
    def __init__(self, progress_bar, log_viewer):
        self.progressBar = progress_bar
        self.log = UiLogger(log_viewer)

    def write_b2b_data(self, stateList, jsonObj, worksheetObj):
        if 'b2b' in jsonObj:
            self.log.info('Writing B2B related data')
            b2bHeadingRow = ['GSTIN of Supplier', 'Invoice Number', 'Invoice Date', 'Invoice Value', 'Place of supply', 'Reverse Charge',
                             'Invoice Type', 'Rate', 'Taxable Value', 'Integrated Tax Paid', 'Central Tax Paid', 'State/UT Tax Paid', 'Cess Paid',
                             'Eligibility For ITC', 'Availed ITC Integrated Tax', 'Availed ITC Central Tax', 'Availed ITC State/UT Tax', 'Availed ITC Cess']
            worksheetObj.append(b2bHeadingRow)
            rowNum = 2
            for b2bRow in jsonObj['b2b']:
                ctin = b2bRow['ctin']
                if 'inv' in b2bRow:
                    for invData in b2bRow['inv']:
                        invNum  = invData['inum']
                        invDate = invData['idt']
                        invVal  = invData['val']
                        invPos  = invData['pos']
                        invRevCharge = invData['rchrg']
                        invType = invData['inv_typ']

                        try:
                            posVal = int(invData['pos'])
                            invPos = '%02d-%s' % (posVal, stateList[posVal - 1])
                        except:
                            invPos = 'State not available'

                        if invType == 'R':
                            invType = 'Regular'

                        if 'itms' in invData:
                            for itemData in invData['itms']:
                                if 'itm_det' in itemData:
                                    userInputCell = 'S%s' % rowNum
                                    igstTaxCell = 'J%s' % rowNum
                                    cgstTaxCell = 'K%s' % rowNum
                                    sgstTaxCell = 'L%s' % rowNum
                                    cessPaidCell = 'M%s' % rowNum
                                    invTaxRate = itemData['itm_det']['rt']
                                    invTaxableVal = itemData['itm_det']['txval']
                                    invIntTaxPaid     = 0
                                    invCentralTaxPaid = 0
                                    invStateTaxPaid   = 0
                                    invCessPaid       = 0
                                    invEligibilityOfItc = 'Inputs'
                                    invAvailedItcIntTax     = '=if(%s="A",%s,0)' % (userInputCell, igstTaxCell)
                                    invAvailedItcCentralTax = '=if(%s="A",%s,0)' % (userInputCell, cgstTaxCell)
                                    invAvailedItcStateTax   = '=if(%s="A",%s,0)' % (userInputCell, sgstTaxCell)
                                    invAvailedItcCess       = '=if(%s="A",%s,0)' % (userInputCell, cessPaidCell)

                                    if 'iamt' in itemData['itm_det']:
                                        invIntTaxPaid = itemData['itm_det']['iamt']

                                    if 'samt' in itemData['itm_det']:
                                        invCentralTaxPaid = itemData['itm_det']['samt']

                                    if 'camt' in itemData['itm_det']:
                                        invStateTaxPaid = itemData['itm_det']['camt']

                                    if invIntTaxPaid != 0 and invCentralTaxPaid != 0 and invStateTaxPaid != 0:
                                        self.log.eror("All three taxes IGST, CGST and SGST present for invoice '%s' dated '%s'" % (invNum, invDate))
                                    else:
                                        b2bDataRow = [ctin, invNum, invDate, invVal, invPos, invRevCharge, invType, invTaxRate, invTaxableVal, invIntTaxPaid, invCentralTaxPaid, invStateTaxPaid, invCessPaid, invEligibilityOfItc, invAvailedItcIntTax, invAvailedItcCentralTax, invAvailedItcStateTax, invAvailedItcCess]
                                        worksheetObj.append(b2bDataRow)
                                        rowNum += 1
                        else:
                            self.log.error("No item data present for invoice '%s' for customer '%s'" % (invNum, ctin))
                else:
                    self.log.error("No Invoice data present for '%s'" % ctin)
        else:
            self.log.warning('Input file does not contain B2B related data')
        return

    def convert(self, json_file_name):
        output_file_name = json_file_name.replace('.json', '.xlsx')
        state_list = commonLib.getStateList()
        # Check input file exists in file system
        self.log.info('Checking input file exists or not')
        if not is_file_exists(json_file_name):
            return

        # Load JSON data from file
        self.log.info("Loading JSON data from '%s'" % json_file_name)
        json_obj = self.loadJsonData(json_file_name)
        if not json_obj:
            self.log.error('Unable to read JSON data from file')
            return

        self.print_gstin(json_obj)

        # Open Excel file and create required work sheets
        workbook = openpyxl.Workbook()
        # hsnWorksheet = workbook.create_sheet('HSN')
        # b2csWorksheet = workbook.create_sheet('B2CS')
        b2b_worksheet = workbook.create_sheet('B2B')
        # errWorksheet = workbook.create_sheet('Error Report')

        try:
            # self.writeHsnData(jsonObj, hsnWorksheet)
            # self.writeB2CSData(stateList, jsonObj, b2csWorksheet)
            self.write_b2b_data(state_list, json_obj, b2b_worksheet)
            # self.writeErrData(stateList, jsonObj, errWorksheet)
        except Exception:
            self.log.error('Exception during file parsing -> %s'
                           % traceback.format_exc())

        workbook.save(output_file_name)
        self.log.info("Output saved in file '%s'" % output_file_name)
        self.log.raw_line(':::::::::: Done ::::::::::')
