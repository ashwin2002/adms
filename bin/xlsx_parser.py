import dbf
import openpyxl
import traceback

import constants
from lib.excel_lib import is_row_blank, is_file_exists, \
                          get_target_excel_class, get_next_col
from lib.logging import UiLogger


class GstrParser:
    def __init__(self, progress_bar, log_viewer):
        self.progressBar = progress_bar
        self.log = UiLogger(log_viewer)

    @staticmethod
    def get_excel_header(sheet_name):
        return getattr(getattr(constants, "headers"), sheet_name)

    @staticmethod
    def get_month_abbr(month_num):
        return ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul',
                'Aug', 'Sep', 'Oct', 'Nov', 'Dec'][month_num]

    def create_dbf(self, parsed_xl_file_name, xl_class):
        start_col = "A"
        row_num = 1
        output_dbf_file_name = "".join(parsed_xl_file_name.split(".")[0:-1]) \
                               + ".dbf"
        self.log.info("Creating DBF file from %s::B2B data"
                      % parsed_xl_file_name)
        b2b_sheet = openpyxl.load_workbook(parsed_xl_file_name)["B2B"]
        fields_in_dbf = [field.strip().split(' ')
                         for field in xl_class.dbf_spec.split(";")]
        fields_in_b2b = list()

        # Read first_row to get the fields from XL
        curr_col = start_col
        null_data = False
        while not null_data:
            cell_val = b2b_sheet[curr_col + str(row_num)].value or ""
            if cell_val.strip() == "":
                break
            fields_in_b2b.append(cell_val)
            curr_col = get_next_col(curr_col)

        dbf_field_len = len(fields_in_dbf)
        b2b_field_len = len(fields_in_b2b)
        self.log.info("Total fields in B2B=%d, dbf_spec=%d"
                      % (b2b_field_len, dbf_field_len))
        if b2b_field_len != dbf_field_len:
            self.log.error("Cannot process in incompatible field count")
            return

        # Create DBF file for writing data
        dbf_table = dbf.Table(output_dbf_file_name,
                              field_specs=xl_class.dbf_spec)
        dbf_table.open(mode=dbf.READ_WRITE)

        # Starting converting Xl -> DBF
        while True:
            curr_col = start_col
            row_num += 1
            row_data = list()
            for index, field_spec in enumerate(fields_in_dbf):
                cell_val = b2b_sheet[curr_col + str(row_num)].value
                if cell_val is None:
                    if index == 0:
                        break
                    else:
                        cell_val = ""
                curr_col = get_next_col(curr_col)

                if field_spec[1][0] == "D":
                    t_val = cell_val.split("-")
                    d_format = "%d"
                    m_format = "%m"
                    y_format = "%Y"
                    if len(t_val[1]) == 3:
                        m_format = "%b"
                    if len(t_val[2]) == 2:
                        y_format = "%y"
                    cell_val = dbf.Date.strptime(cell_val,
                                                 "%s-%s-%s" % (d_format,
                                                               m_format,
                                                               y_format))
                elif field_spec[1][0] == "N":
                    if cell_val == "":
                        cell_val = 0
                elif field_spec[1][0] == "C":
                    cell_val = str(cell_val)[:int(field_spec[1][2:-1])]

                row_data.append(cell_val)
            if not row_data:
                break

            dbf_table.append(tuple(row_data))

        dbf_table.close()
        self.log.info("Done writing DBF file %s" % output_dbf_file_name)

    def shuffle_row_for_b2b_headers(self, input_sheet_name, rows_to_process,
                                    dwnload_val, gst_2yrm_val, xl_class):
        def get_var_name(header_dict, val):
            for tem_header, t_val in header_dict.items():
                if t_val == val:
                    return tem_header
            return None

        index_map = list()
        rows_to_append = list()
        b2b_headers_class = constants.headers.B2B
        output_header = self.get_excel_header(input_sheet_name)

        for b2b_header in xl_class.output_header["B2B"]:
            b2b_header_var = get_var_name(b2b_headers_class.get_dict(),
                                          b2b_header)
            index = -1
            for t_index, t_header in enumerate(
                    xl_class.output_header[input_sheet_name]):
                target_header_var = get_var_name(output_header.get_dict(),
                                                 t_header)
                if b2b_header_var == target_header_var:
                    index = t_index
                    break
            index_map.append(index)

        for input_row in rows_to_process:
            row_data = list()
            for index in index_map:
                if index == -1:
                    row_data.append("")
                    continue
                row_data.append(input_row[index])
            row_data[xl_class.output_header["B2B"].index(
                b2b_headers_class.DOWNLOAD)] = dwnload_val
            row_data[xl_class.output_header["B2B"].index(
                b2b_headers_class.GST2_YRM)] = gst_2yrm_val

            d_index = xl_class.output_header["B2B"].index(
                b2b_headers_class.INV_TYPE)
            if input_sheet_name != "B2B" and row_data[d_index] == "":
                row_data.pop(d_index)
                row_data.insert(d_index, input_sheet_name)
            if row_data[d_index].lower() == "credit note":
                for value_header in [b2b_headers_class.INV_VALUE,
                                     b2b_headers_class.TAXABLE,
                                     b2b_headers_class.IGST_PAID,
                                     b2b_headers_class.CGST_PAID,
                                     b2b_headers_class.SGST_PAID,
                                     b2b_headers_class.CESS_PAID]:
                    header_index = xl_class.output_header["B2B"].index(
                        value_header)
                    value = row_data.pop(header_index)
                    row_data.insert(header_index, -value)
            rows_to_append.append(row_data)
        return rows_to_append


class Gstr2Parser(GstrParser):
    def __init__(self, progress_bar, log_viewer):
        super(Gstr2Parser, self).__init__(progress_bar, log_viewer)

    def extract_data_from_excel_sheet(self, input_sheet_name, input_sheet,
                                      row_num, dwnload_val, gst_2yrm_val,
                                      xl_class):
        rows_to_append = list()
        num_of_blank_rows_to_break = 5
        num_blank_rows = 0

        sheet_header = xl_class.output_header[input_sheet_name]
        input_header = xl_class.input_header[input_sheet_name]

        t_header = self.get_excel_header(input_sheet_name)
        gstin_index = inv_num_index = -1
        if t_header.GSTIN in input_header:
            gstin_index = input_header.index(t_header.GSTIN)
        if t_header.INV_NO in input_header:
            inv_num_index = input_header.index(t_header.INV_NO)
        while True:
            if num_blank_rows == num_of_blank_rows_to_break:
                break

            data_row = [None for _ in sheet_header]
            input_row = list()

            curr_col = "A"
            for index in range(len(sheet_header)):
                cell_val = input_sheet[curr_col + str(row_num)].value
                input_row.append(cell_val)
                curr_col = get_next_col(curr_col)

            row_num += 1
            if is_row_blank(input_row):
                num_blank_rows += 1
                continue

            num_blank_rows = 0
            if gstin_index != -1:
                gstin_val = input_row[gstin_index]
                if gstin_val is None or len(gstin_val) != 15 \
                        or str(input_row[inv_num_index]).find("-Total") != -1:
                    continue

            if inv_num_index != -1:
                # input_row[inv_num_index] = "=CONCATENATE(\"%s\")" \
                #                            % input_row[inv_num_index]
                input_row[inv_num_index] = input_row[inv_num_index]
            for index, d_type in enumerate(sheet_header):
                if d_type == t_header.DUMMY:
                    continue
                try:
                    d_index = input_header.index(d_type)
                    val = input_row[d_index]
                    if d_type == t_header.SUBMITTED:
                        val = "SUBMITTED" if val == "Y" else "NOT_SUBMITTED"
                    data_row[index] = val
                except ValueError:
                    pass

            if t_header.INV_TYPE in sheet_header and data_row[
                    sheet_header.index(t_header.INV_TYPE)] == "R":
                data_row[sheet_header.index(t_header.INV_TYPE)] = "Regular"

            if t_header.INV_DATE in sheet_header:
                cell_val = \
                    data_row[sheet_header.index(t_header.INV_DATE)].split("-")
                cell_val.insert(1, self.get_month_abbr(int(cell_val.pop(1))))
                data_row[sheet_header.index(t_header.INV_DATE)] = \
                    "-".join(cell_val)

            if input_sheet_name == "B2B":
                d_index = sheet_header.index(t_header.ELIGIBLE)
                data_row.pop(d_index)
                data_row.insert(d_index, "Inputs")
                for d_type in [t_header.IGST_AVAILED,
                               t_header.CGST_AVAILED,
                               t_header.SGST_AVAILED,
                               t_header.CESS_AVAILED]:
                    d_index = sheet_header.index(d_type)
                    data_row.pop(d_index)
                    data_row.insert(d_index, 0.00)
                data_row[sheet_header.index(
                    t_header.DOWNLOAD)] = dwnload_val
                data_row[sheet_header.index(
                    t_header.GST2_YRM)] = gst_2yrm_val
            elif input_sheet_name == "IMPG":
                data_row[sheet_header.index(t_header.GSTIN)] = \
                    "07INDIANCUSTOMS"
                data_row[sheet_header.index(t_header.SUPP_NAME)] = \
                    "Indian Customs"
                data_row[sheet_header.index(t_header.INV_VALUE)] = \
                    data_row[sheet_header.index(t_header.TAXABLE)] \
                    + data_row[sheet_header.index(t_header.IGST_PAID)]
                data_row[sheet_header.index(t_header.SUBMITTED)] = 'SUBMITTED'
                data_row[sheet_header.index(t_header.GST3_YRM)] = 'Y'
            rows_to_append.append(data_row)
        return rows_to_append

    def convert(self, input_file_name):
        self.log.info("Processing file '%s'" % input_file_name)
        output_file_name = input_file_name.split('/')
        file_name = (output_file_name.pop()).split('.')[0]
        output_file_name = '/'.join(output_file_name)
        output_file_name += '/%s_parsed.xlsx' % file_name

        file_name_data = \
            input_file_name \
            .split("/")[-1] \
            .split('.')[0] \
            .split('_')
        gst_2yrm_val = file_name_data[1][2:] + file_name_data[1][0:2]
        dwnload_val = \
            str(file_name_data[3][0:2]) + '-' \
            + str(file_name_data[3][2:5]) + '-' \
            + str(file_name_data[3][5:])

        xl_class = get_target_excel_class(gst_2yrm_val)
        try:
            # Load input / output workbooks
            input_workbook = openpyxl.load_workbook(input_file_name)
            if not is_file_exists(output_file_name):
                output_workbook = openpyxl.Workbook()
                output_workbook.guess_types = True
            else:
                output_workbook = openpyxl.load_workbook(input_file_name)

            output_workbook_sheets = output_workbook.get_sheet_names()

            for input_sheet_name in xl_class.input_header.keys():
                if input_sheet_name not in input_workbook.get_sheet_names():
                    continue

                self.log.info('Processing sheet %s' % input_sheet_name)
                if input_sheet_name in output_workbook_sheets:
                    self.log.warning('Sheet %s already exists! Will recreate..'
                                     % input_sheet_name)
                    output_workbook.remove(
                        output_workbook.get_sheet_by_name(input_sheet_name))
                input_sheet = input_workbook.get_sheet_by_name(
                    input_sheet_name)

                # Start of data extraction logic
                rows_to_append = self.extract_data_from_excel_sheet(
                    input_sheet_name, input_sheet,
                    xl_class.start_row[input_sheet_name],
                    dwnload_val, gst_2yrm_val, xl_class)

                # Create and append data to output sheet
                output_sheet = output_workbook.create_sheet(input_sheet_name)
                output_sheet.append(
                    xl_class.output_header[input_sheet_name])
                for row in rows_to_append:
                    output_sheet.append(row)

                # Logic to append data to B2B sheet
                if input_sheet_name != "B2B":
                    rows_to_append = self.shuffle_row_for_b2b_headers(
                        input_sheet_name,
                        rows_to_append,
                        dwnload_val,
                        gst_2yrm_val,
                        xl_class)
                    output_sheet = output_workbook["B2B"]

                    for row in rows_to_append:
                        output_sheet.append(row)

            output_workbook.save(output_file_name)
            self.log.info("Output saved in file '%s'" % output_file_name)
            self.create_dbf(output_file_name, xl_class)
        except Exception:
            self.log.error('Exception during file parsing -> %s'
                           % (traceback.format_exc()))

        self.log.raw_line(':::::::::: Done ::::::::::')
